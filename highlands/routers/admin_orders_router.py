"""Admin order management endpoints."""
import html as _html
import urllib.parse
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from highlands.database import get_db
from highlands import models
from highlands.auth_utils import require_admin
import io

router = APIRouter(prefix="/api/admin/orders", tags=["admin-orders"])


class OrderItemOut(BaseModel):
    id: int
    product_id: int | None
    name: str
    price: int
    quantity: int
    subtotal: int

    class Config:
        from_attributes = True


class OrderOut(BaseModel):
    id: int
    customer_name: str
    phone: str
    address: str | None
    total: int
    note: str | None
    status: str
    payment_status: str
    payment_method: str
    created_at: str | None
    user_id: int | None
    items: list[OrderItemOut]

    class Config:
        from_attributes = True


class OrderStatusUpdate(BaseModel):
    status: str | None = None
    payment_status: str | None = None


class OrderItemCreate(BaseModel):
    product_id: int
    quantity: int


class OrderCreate(BaseModel):
    customer_name: str = "Khách lẻ"
    phone: str = "—"
    address: str | None = None
    note: str | None = None
    payment_method: str = "cash"   # cash / qr_transfer
    items: list[OrderItemCreate]


class PaginatedOrders(BaseModel):
    items: list[OrderOut]
    total: int
    skip: int
    limit: int
    has_next: bool


def _build_order_query(
    db: Session,
    status: str | None = None,
    search: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    min_total: int | None = None,
    max_total: int | None = None,
    month: int | None = None,
    year: int | None = None,
):
    from datetime import datetime
    from sqlalchemy import extract

    query = db.query(models.Order).filter(models.Order.is_active == 1)

    if status:
        query = query.filter(models.Order.status == status)

    if search:
        query = query.filter(
            models.Order.customer_name.ilike(f"%{search}%")
            | models.Order.phone.ilike(f"%{search}%")
        )

    if year:
        query = query.filter(extract("year", models.Order.created_at) == year)

    if month:
        query = query.filter(extract("month", models.Order.created_at) == month)

    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
            query = query.filter(models.Order.created_at >= start)
        except Exception:
            pass

    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            query = query.filter(models.Order.created_at <= end)
        except Exception:
            pass

    if min_total is not None:
        query = query.filter(models.Order.total >= min_total)

    if max_total is not None:
        query = query.filter(models.Order.total <= max_total)

    return query


@router.get("/export")
def export_orders_excel(
    status: str | None = None,
    search: str | None = None,
    month: int | None = None,
    year: int | None = None,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export filtered orders to Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    query = _build_order_query(db, status=status, search=search, month=month, year=year)
    orders = query.order_by(models.Order.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active

    label_parts = []
    if month and year:
        label_parts.append(f"Tháng {month:02d}/{year}")
    elif month:
        label_parts.append(f"Tháng {month:02d}")
    elif year:
        label_parts.append(f"Năm {year}")
    if status:
        status_vn = {"pending": "Đang Chờ", "confirmed": "Xác Nhận", "done": "Hoàn Thành"}.get(status, status)
        label_parts.append(status_vn)

    ws.title = "Đơn Hàng"
    title_text = "DANH SÁCH ĐƠN HÀNG"
    if label_parts:
        title_text += " - " + " | ".join(label_parts)

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="8B4513")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Summary row
    total_revenue = sum(o.total for o in orders)
    ws.merge_cells("A2:J2")
    summary_cell = ws["A2"]
    summary_cell.value = f"Tổng đơn hàng: {len(orders)}  |  Tổng doanh thu: {total_revenue:,}đ".replace(",", ".")
    summary_cell.font = Font(italic=True, size=11, color="5D4037")
    summary_cell.fill = PatternFill(fill_type="solid", fgColor="FFF3E0")
    summary_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Header
    headers = ["#", "Mã ĐH", "Khách Hàng", "Số Điện Thoại", "Địa Chỉ", "Tổng Tiền (đ)", "Trạng Thái", "Thanh Toán", "Ghi Chú", "Ngày Tạo"]
    header_fill = PatternFill(fill_type="solid", fgColor="D7CCC8")
    thin = Side(style="thin", color="BCAAA4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    status_vn_map = {"pending": "Đang Chờ", "confirmed": "Xác Nhận", "done": "Hoàn Thành"}
    payment_vn_map = {"paid": "Đã Thanh Toán", "unpaid": "Chưa Thanh Toán"}

    row_fills = [PatternFill(fill_type="solid", fgColor="FFFFFF"), PatternFill(fill_type="solid", fgColor="FBF9F7")]

    for i, order in enumerate(orders, start=1):
        row_num = i + 3
        fill = row_fills[i % 2]
        values = [
            i,
            f"#{order.id}",
            order.customer_name,
            order.phone,
            order.address or "",
            order.total,
            status_vn_map.get(order.status, order.status),
            payment_vn_map.get(getattr(order, "payment_status", "unpaid"), "Chưa TT"),
            order.note or "",
            order.created_at.strftime("%d/%m/%Y %H:%M") if order.created_at else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = [5, 9, 22, 16, 30, 18, 14, 18, 22, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header rows
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_parts = ["don_hang"]
    if year:
        filename_parts.append(str(year))
    if month:
        filename_parts.append(f"thang{month:02d}")
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("", response_model=PaginatedOrders)
def list_orders(
    skip: int = 0,
    limit: int = 20,
    status: str | None = None,
    search: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    min_total: int | None = None,
    max_total: int | None = None,
    month: int | None = None,
    year: int | None = None,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    query = _build_order_query(
        db,
        status=status,
        search=search,
        start_date=start_date,
        end_date=end_date,
        min_total=min_total,
        max_total=max_total,
        month=month,
        year=year,
    )

    total = query.count()
    db_items = query.order_by(models.Order.created_at.desc()).offset(skip).limit(limit).all()

    items = []
    for order in db_items:
        items.append({
            "id": order.id,
            "customer_name": order.customer_name,
            "phone": order.phone,
            "address": order.address,
            "total": order.total,
            "note": order.note,
            "status": order.status,
            "payment_status": order.payment_status or "unpaid",
            "payment_method": order.payment_method or "cash",
            "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
            "user_id": order.user_id,
            "items": [
                {
                    "id": item.id,
                    "product_id": item.product_id,
                    "name": item.name,
                    "price": item.price,
                    "quantity": item.quantity,
                    "subtotal": item.subtotal,
                }
                for item in (order.items or [])
            ],
        })

    return {
        "items": items,
        "total": total,
        "skip": skip,
        "limit": limit,
        "has_next": skip + limit < total,
    }


@router.get("/{order_id}", response_model=OrderOut)
def get_order(
    order_id: int,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Not found")

    return {
        "id": order.id,
        "customer_name": order.customer_name,
        "phone": order.phone,
        "address": order.address,
        "total": order.total,
        "note": order.note,
        "status": order.status,
        "payment_status": order.payment_status or "unpaid",
        "payment_method": order.payment_method or "cash",
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
        "user_id": order.user_id,
        "items": [
            {
                "id": item.id,
                "product_id": item.product_id,
                "name": item.name,
                "price": item.price,
                "quantity": item.quantity,
                "subtotal": item.subtotal,
            }
            for item in (order.items or [])
        ],
    }


@router.patch("/{order_id}", response_model=OrderOut)
def update_order_status(
    order_id: int,
    body: OrderStatusUpdate,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Not found")

    if body.status is not None:
        valid_statuses = ["pending", "confirmed", "done"]
        if body.status not in valid_statuses:
            raise HTTPException(status_code=400, detail="Invalid status")
        order.status = body.status

    if body.payment_status is not None:
        valid_payment_statuses = ["unpaid", "paid"]
        if body.payment_status not in valid_payment_statuses:
            raise HTTPException(status_code=400, detail="Invalid payment_status")
        order.payment_status = body.payment_status

    db.commit()
    db.refresh(order)

    return {
        "id": order.id,
        "customer_name": order.customer_name,
        "phone": order.phone,
        "address": order.address,
        "total": order.total,
        "note": order.note,
        "status": order.status,
        "payment_status": order.payment_status or "unpaid",
        "payment_method": order.payment_method or "cash",
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
        "user_id": order.user_id,
        "items": [
            {
                "id": item.id,
                "product_id": item.product_id,
                "name": item.name,
                "price": item.price,
                "quantity": item.quantity,
                "subtotal": item.subtotal,
            }
            for item in (order.items or [])
        ],
    }


@router.post("", status_code=201, response_model=OrderOut)
def create_order(
    body: OrderCreate,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Create a new order with items."""
    if not body.items:
        raise HTTPException(status_code=400, detail="Order must have at least one item")

    total = 0
    order_items = []

    for item in body.items:
        product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item.product_id} not found")

        subtotal = product.price * item.quantity
        total += subtotal

        order_items.append({
            "product_id": product.id,
            "name": product.name,
            "price": product.price,
            "quantity": item.quantity,
            "subtotal": subtotal,
        })

    order = models.Order(
        customer_name=body.customer_name,
        phone=body.phone,
        address=body.address,
        total=total,
        note=body.note,
        status="pending",
        payment_method=body.payment_method,
        is_active=1,
    )
    db.add(order)
    db.flush()

    for item_data in order_items:
        order_item = models.OrderItem(
            order_id=order.id,
            **item_data,
        )
        db.add(order_item)

    db.commit()
    db.refresh(order)

    return {
        "id": order.id,
        "customer_name": order.customer_name,
        "phone": order.phone,
        "address": order.address,
        "total": order.total,
        "note": order.note,
        "status": order.status,
        "payment_status": order.payment_status or "unpaid",
        "payment_method": order.payment_method or "cash",
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else None,
        "user_id": order.user_id,
        "items": [
            {
                "id": item.id,
                "product_id": item.product_id,
                "name": item.name,
                "price": item.price,
                "quantity": item.quantity,
                "subtotal": item.subtotal,
            }
            for item in (order.items or [])
        ],
    }


def _build_invoice_html(order: models.Order, items: list) -> str:
    from highlands.config import BANK_CODE, BANK_ACCOUNT, BANK_ACCOUNT_NAME, SHOP_NAME, SHOP_ADDRESS

    shop_name = _html.escape(SHOP_NAME or "Tu's Coffee")
    shop_address = _html.escape(SHOP_ADDRESS or "")
    customer_name = _html.escape(order.customer_name or "")
    phone = _html.escape(order.phone or "")
    address = _html.escape(order.address or "")
    note = _html.escape(order.note or "")
    created_str = order.created_at.strftime("%H:%M %d/%m/%Y") if order.created_at else "—"

    def fmt(n: int) -> str:
        return f"{n:,}".replace(",", ".") + "đ"

    rows = ""
    for i, item in enumerate(items):
        bg = "#fafafa" if i % 2 == 0 else "#fff"
        rows += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 8px;">{_html.escape(item.name or "")}</td>'
            f'<td style="text-align:center;padding:6px 4px;">{item.quantity}</td>'
            f'<td style="text-align:right;padding:6px 8px;">{fmt(item.price)}</td>'
            f'<td style="text-align:right;padding:6px 8px;font-weight:600;">{fmt(item.subtotal)}</td>'
            f"</tr>"
        )

    pm_label = "💵 Tiền mặt" if order.payment_method == "cash" else "📱 Chuyển khoản QR"
    ps_label = "✅ Đã thanh toán" if order.payment_status == "paid" else "⏳ Chưa thanh toán"

    addr_html = f'<div class="shop-sub">{shop_address}</div>' if shop_address else ""
    address_html = (
        f'<div style="grid-column:1/-1"><span class="lbl">Địa chỉ:</span> {address}</div>'
        if address else ""
    )
    note_html = (
        f'<div style="grid-column:1/-1"><span class="lbl">Ghi chú:</span> {note}</div>'
        if note else ""
    )

    qr_section = ""
    if BANK_ACCOUNT and order.payment_status != "paid":
        acct_enc = urllib.parse.quote(BANK_ACCOUNT_NAME or SHOP_NAME or "")
        qr_url = (
            f"https://img.vietqr.io/image/{BANK_CODE}-{BANK_ACCOUNT}-compact2.png"
            f"?amount={order.total}&addInfo=DH{order.id}&accountName={acct_enc}"
        )
        bk = _html.escape(BANK_CODE or "")
        acct = _html.escape(BANK_ACCOUNT or "")
        qr_section = (
            f'<div style="margin-top:16px;text-align:center;">'
            f'<div style="font-size:12px;font-weight:600;color:#555;margin-bottom:10px;">📱 Quét mã QR để thanh toán</div>'
            f'<img src="{qr_url}" alt="QR Code" style="width:180px;height:180px;border-radius:8px;border:1px solid #eee;" '
            f'onerror="this.parentElement.innerHTML=\'<p style=&quot;color:#999;font-size:12px&quot;>Không tải được QR. Kiểm tra kết nối mạng.</p>\'">'
            f'<div style="margin-top:10px;font-size:11px;color:#777;line-height:1.9;">'
            f"<div>Ngân hàng: <strong>{bk}</strong></div>"
            f"<div>Số tài khoản: <strong>{acct}</strong></div>"
            f"<div>Số tiền: <strong>{fmt(order.total)}</strong></div>"
            f"<div>Nội dung CK: <strong>DH{order.id}</strong></div>"
            f"</div></div>"
        )

    css = (
        "* { margin:0; padding:0; box-sizing:border-box; }"
        "body { font-family:'Inter',sans-serif; max-width:420px; margin:0 auto; padding:20px; font-size:13px; color:#333; background:#fff; }"
        ".shop-name { font-size:22px; font-weight:700; color:#C8102E; text-align:center; }"
        ".shop-sub { font-size:11px; color:#888; text-align:center; margin-top:3px; }"
        ".divider { border:none; border-top:1px dashed #ddd; margin:12px 0; }"
        ".info-grid { display:grid; grid-template-columns:1fr 1fr; gap:5px; font-size:12px; margin:8px 0; }"
        ".lbl { color:#888; }"
        "table { width:100%; border-collapse:collapse; font-size:12px; }"
        "th { background:#fef2f2; color:#C8102E; padding:7px 8px; text-align:left; font-size:11px; font-weight:600; }"
        ".total-row { display:flex; justify-content:space-between; font-size:15px; font-weight:700; color:#C8102E; margin-top:8px; }"
        ".status-row { display:flex; justify-content:space-between; font-size:12px; color:#666; margin-top:4px; }"
        ".footer { text-align:center; color:#aaa; font-size:11px; margin-top:16px; }"
        ".print-btn { display:block; margin:20px auto 0; padding:10px 28px; background:#C8102E; color:#fff; border:none; border-radius:8px; cursor:pointer; font-size:14px; font-weight:600; font-family:inherit; }"
        ".print-btn:hover { background:#9B0C22; }"
        "@media print { .no-print { display:none!important; } body { padding:10px; max-width:100%; } }"
    )

    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width,initial-scale=1.0">
    <title>Hóa Đơn #{order.id}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>{css}</style>
</head>
<body>
    <div class="shop-name">{shop_name}</div>
    {addr_html}
    <div class="shop-sub" style="margin-top:2px;">HÓA ĐƠN BÁN HÀNG</div>
    <hr class="divider">
    <div class="info-grid">
        <div><span class="lbl">Mã đơn:</span> <strong>#{order.id}</strong></div>
        <div><span class="lbl">Ngày:</span> {created_str}</div>
        <div><span class="lbl">Khách hàng:</span> {customer_name}</div>
        <div><span class="lbl">SĐT:</span> {phone}</div>
        {address_html}
        {note_html}
    </div>
    <hr class="divider">
    <table>
        <thead>
            <tr>
                <th>Sản phẩm</th>
                <th style="text-align:center">SL</th>
                <th style="text-align:right">Đơn giá</th>
                <th style="text-align:right">Thành tiền</th>
            </tr>
        </thead>
        <tbody>{rows}</tbody>
    </table>
    <hr class="divider">
    <div class="total-row"><span>TỔNG CỘNG</span><span>{fmt(order.total)}</span></div>
    <div class="status-row"><span>{pm_label}</span><span>{ps_label}</span></div>
    {qr_section}
    <hr class="divider">
    <div class="footer">☕ Cảm ơn quý khách! Hẹn gặp lại.</div>
    <button class="print-btn no-print" onclick="window.print()">🖨️ In Hóa Đơn</button>
    <script>window.addEventListener('load', () => setTimeout(() => window.print(), 1000));</script>
</body>
</html>"""


@router.get("/{order_id}/invoice", response_class=HTMLResponse)
def get_order_invoice(
    order_id: int,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return a printable HTML invoice for an order."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng")
    return HTMLResponse(content=_build_invoice_html(order, order.items or []))


@router.get("/{order_id}/reviews")
def get_order_reviews(
    order_id: int,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get all reviews by the order's user for products in this order."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Not found")

    product_ids = [item.product_id for item in (order.items or []) if item.product_id]

    if not product_ids or not order.user_id:
        return []

    reviews = (
        db.query(models.ProductReview)
        .filter(
            models.ProductReview.product_id.in_(product_ids),
            models.ProductReview.user_id == order.user_id,
        )
        .all()
    )

    return [
        {
            "id": r.id,
            "product_id": r.product_id,
            "product_name": r.product.name if r.product else "—",
            "stars": r.stars,
            "comment": r.comment,
            "created_at": r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else None,
        }
        for r in reviews
    ]


@router.delete("/reviews/{review_id}")
def delete_review(
    review_id: int,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete a product review."""
    review = db.query(models.ProductReview).filter(models.ProductReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Đánh giá không tồn tại")
    db.delete(review)
    db.commit()
    return {"detail": "Đã xóa đánh giá"}


@router.delete("/{order_id}")
def delete_order(
    order_id: int,
    admin: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Soft delete an order (set is_active=0)."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.is_active = 0
    db.commit()

    return {"success": True, "message": "Order deleted successfully"}
